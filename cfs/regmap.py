import re
from openpyxl import load_workbook

class RegMap():
    def __init__(self):
        pass

    def __init_data(self, data):
        sregex = re.compile(r'(\d+.\d+.\d+)\sAddr\s(0x[0-9a-fA-F]{8})\s(.+)')
        mo = sregex.search(data)
        regdata = {}
        regdata['addr'] = int(mo.group(2), 16)
        regdata['description'] = mo.group(3)
        regdata['value'] = 0
        return regdata

    def load(self, worksheet):
        self.__regmaps = load_workbook(worksheet, data_only=True)

    def get_regblocks(self):
        return self.__regmaps.sheetnames[2:]

    def get_regmap(self, block):
        regmap = self.__regmaps[block]
        regmaplist = [ 
        self.__init_data(d[0]) 
        for d in regmap.iter_rows(min_row=1, max_row=regmap.max_row, min_col=3, max_col=3, values_only=True) 
        if None not in d and 'Addr' in d[0] ]
        return regmaplist 
